import streamlit as st
import openpyxl
import os
from datetime import date
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

archivo_excel = 'calibraciones.xlsx'

# Inicializar Excel
def inicializar_excel():
    if not os.path.exists(archivo_excel):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Orden de Servicio', 'Fecha', 'Equipo', 'Certificado', 'Cliente', 'Sede/Servicio',
                   'Conformidad', 'Ejecutado por', 'Firmado por', 'Avalado por'])
        wb.save(archivo_excel)
    else:
        wb = openpyxl.load_workbook(archivo_excel)
        ws = wb.active
        if ws.max_row < 2:
            ws.append(['Orden de Servicio', 'Fecha', 'Equipo', 'Certificado', 'Cliente', 'Sede/Servicio',
                       'Conformidad', 'Ejecutado por', 'Firmado por', 'Avalado por'])
            wb.save(archivo_excel)

# Determinar letra del equipo
def determinar_letra_equipo(equipo):
    equipos_E = ["Tensiometro Digital", "Tensiometro Adulto", "Tensiometro Pediátrico"]
    equipos_B = ["Balanza Adulto", "Pesa Bebé", "Gramera", "Dinamómetro"]
    if equipo in equipos_E:
        return 'E'
    elif equipo in equipos_B:
        return 'B'
    else:
        return 'X'

# Subir archivo a Google Drive usando OAuth 2.0 de Usuario
def subir_a_drive_oauth(nombre_archivo):
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    
    flow = InstalledAppFlow.from_client_config(st.secrets["google_oauth"], SCOPES)    
    creds = flow.run_local_server(port=0)

    service = build('drive', 'v3', credentials=creds)

    file_metadata = {
        'name': nombre_archivo
    }

    media = MediaFileUpload(nombre_archivo, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    st.success(f"Archivo subido a Google Drive. ID: {file.get('id')}")

# Inicialización
inicializar_excel()

# Interfaz de usuario en Streamlit
st.title("Asignación de Stickers de Calibración")

# NUEVA CASILLA: Número Consecutivo Inicial Manual
consecutivo_manual = st.number_input("Número de Consecutivo Inicial (Manual)", min_value=1, step=1, value=1588)

modo_asignacion = st.checkbox("Modo Asignación de Stickers (simplificado)")

orden = st.text_input("N° Orden Servicio")
fecha = st.date_input("Fecha", value=date.today())
cliente = st.text_input("Cliente")

if not modo_asignacion:
    sede = st.text_input("Sede o Servicio")
    conformidad = st.selectbox("Conformidad", ["Pasa", "No pasa"])
    ejecutado = st.selectbox("Ejecutado por", ["Edwin Garzon", "Juan Melo", "Vratislav Cala"])

st.markdown("---")

st.subheader("Agregar Equipos")
equipo = st.selectbox("Tipo de Equipo", [
    "Tensiometro Digital",
    "Tensiometro Adulto",
    "Tensiometro Pediátrico",
    "Balanza Adulto",
    "Pesa Bebé",
    "Gramera",
    "Dinamómetro"
])
cantidad = st.number_input("Cantidad", min_value=1, step=1)

if 'equipos' not in st.session_state:
    st.session_state.equipos = []

if st.button("Agregar Equipo"):
    st.session_state.equipos.append((equipo, cantidad))

st.write("### Equipos añadidos:")
for eq, cant in st.session_state.equipos:
    st.write(f"{cant} x {eq}")

if st.button("Guardar y Subir a Google Drive"):
    if not st.session_state.equipos:
        st.error("Debes agregar al menos un equipo")
    elif not orden or not cliente:
        st.error("Orden de Servicio y Cliente son obligatorios")
    else:
        wb = openpyxl.load_workbook(archivo_excel)
        ws = wb.active

        # Limpiar filas desde la fila 2
        ws.delete_rows(2, ws.max_row)

        certificados_asignados = []
        letras_equipos = {'E': [], 'B': []}

        for equipo, cantidad in st.session_state.equipos:
            letra = determinar_letra_equipo(equipo)
            if letra == 'X':
                st.error(f"Equipo inválido: {equipo}")
                st.stop()
            letras_equipos[letra].append((equipo, cantidad))

        # USAR EL CONSECUTIVO MANUAL COMO BASE
        consecutivo_actual = consecutivo_manual

        for letra, lista in letras_equipos.items():
            if not lista:
                continue
            for equipo, cantidad in lista:
                for i in range(cantidad):
                    nuevo_num = consecutivo_actual
                    cert_num = f"{letra}-25-{nuevo_num}"
                    ws.append([
                        orden,
                        fecha.strftime('%Y-%m-%d'),
                        equipo,
                        cert_num,
                        cliente,
                        sede if not modo_asignacion else '',
                        conformidad if not modo_asignacion else '',
                        ejecutado if not modo_asignacion else '',
                        "Angie Rodriguez" if not modo_asignacion else '',
                        "Vratislav Cala" if not modo_asignacion else ''
                    ])
                    certificados_asignados.append(cert_num)
                    consecutivo_actual += 1

        wb.save(archivo_excel)
        wb.close()
        st.success(f"Certificados asignados: {', '.join(certificados_asignados)}")
        subir_a_drive_oauth(archivo_excel)
        st.session_state.equipos = []

# BOTÓN DE DESCARGA DIRECTA DEL EXCEL
with open(archivo_excel, "rb") as file:
    btn = st.download_button(
        label="📥 Descargar Excel",
        data=file,
        file_name=archivo_excel,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
